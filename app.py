import asyncio
import os
import time
import shutil
import openpyxl
import re
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from ttkbootstrap.dialogs import Messagebox

ruta = "utils/codigosSIM.txt"
ruta_alta = "PMC_Creados/"

class App(ttk.Window):
    def __init__(self, root):
        self.root = root
        self.init_ui()
        self.createWidget()

    def init_ui(self):
        self.root.resizable(False, False)
        self.root.iconbitmap("utils/logoOSE.ico")
        self.root.title("Alta Punto SCADA")
        self.root.geometry("600x620")

    def createWidget(self):
        main_frame = ttk.Frame(self.root, padding=20)
        main_frame.pack(fill=BOTH, expand=True)
        
        #titulo
        title_label = ttk.Label(main_frame, text="Alta Punto SCADA", font=("Arial", 20))
        title_label.pack(pady=20)
        
        #infomarcion de Tabla de Variables
        ttk.Label(main_frame, text="Tablas de Variables conforme version 1.0.7 del Programa control SGD-RTUX", font=("Arial", 8), bootstyle="secondary").pack(pady=(0,10))
        
        #frame para los inputs
        input_frame = ttk.Labelframe(main_frame, text="Datos del PMC", padding=15)
        input_frame.pack(pady=(0,20), fill=X)
        
        #Input para el nombre del PMC
        name_label = ttk.Label(input_frame, text="Nombre del PMC")
        name_label.pack(pady=(0,10))
        self.pmc_name_var = ttk.StringVar()
        self.pmc_name_entry = ttk.Entry(input_frame, textvariable=self.pmc_name_var)
        self.pmc_name_entry.pack(pady=(0,10))
        
        #Input para el Codigo del PMC
        code_label = ttk.Label(input_frame, text="Codigo del PMC")
        code_label.pack(pady=(0,10))
        self.pmc_code_var = ttk.StringVar()
        self.pmc_code_entry = ttk.Entry(input_frame, textvariable=self.pmc_code_var)
        self.pmc_code_entry.pack(pady=(0,10))
        
        #Radio buttons para seleccionar el tipo de Hardware
        hardware_label = ttk.Label(input_frame, text="Tipo de Hardware")
        hardware_label.pack(pady=(0,20))
        self.hardware_var = ttk.StringVar()
        hardware_options = ["RTU+","RTUX"]
        for option in hardware_options:
            ttk.Radiobutton(input_frame, text=option, variable=self.hardware_var, value=option).pack(pady=(0,10))
        
        # SOLUCIÓN: Llamar correctamente la función asíncrona
        create_button = ttk.Button(
            main_frame, 
            text="Crear PMC File", 
            bootstyle="success",
            command=self.handle_create_pmc  # ← Cambio aquí
        )
        create_button.pack(pady=30, fill=X)
        #footer con separador y texto de desarrollador
        ttk.Separator(main_frame, bootstyle="secondary").pack(pady=5, fill=X)
        ttk.Label(main_frame, text="Desarrollado por: Marcelo Rodriguez SGD - 2026", font=("Arial", 10), bootstyle="secondary").pack(pady=10)
        ttk.Label(main_frame, text="Ver 1.0", font=("Arial", 9), bootstyle="secondary").pack(pady=(0,10))

    def handle_create_pmc(self):
        """
        Wrapper síncrono para llamar a la función asíncrona
        """
        asyncio.run(self.createPMCFile())  # ← Ahora sí con ()

    def validateInputs(self):
        """Valida que los inputs no esten vacios"""
        if not self.pmc_name_var.get():
            Messagebox.show_error("El nombre del PMC no puede estar vacio","Error")
            return False
        if not self.pmc_code_var.get():
            Messagebox.show_error("El codigo del PMC no puede estar vacio","Error")
            return False
        if not self.hardware_var.get():
            Messagebox.show_error("Debe seleccionar un tipo de hardware","Error")
            return False
        return True

    async def createPMCFile(self):
        """Crea el archivo PMC con los datos ingresados"""
        if not self.validateInputs():
            return
        
        pmc_name = self.pmc_name_var.get().capitalize()
        pmc_code = self.pmc_code_var.get().capitalize()
        hardware_type = self.hardware_var.get()

        answer = Messagebox.yesno(
            f"PMC File - Confirmar datos\nNombre: {pmc_name}\nCodigo: {pmc_code}\nHardware: {hardware_type}", 
            "PMC File"
        )  

        if answer != "Yes":
            return
            
        # Crear carpeta
        createPMCFolder(pmc_name)
        
        # Buscar códigos
        codigosPMC = searchPMCCode(ruta)
        if not codigosPMC:
            return
        
        try:
            # Ejecutar funciones asíncronas
            await createTGDFile(pmc_code)
            await createCSVFile(hardware_type, pmc_name, pmc_code, codigosPMC)
            
            Messagebox.ok("Archivo PMC creado con éxito", "PMC File")
            
            # Limpiar inputs
            self.pmc_code_var.set("")
            self.pmc_name_var.set("")
            self.hardware_var.set("")
            
        except Exception as e:
            Messagebox.show_error(f"Error al crear los archivos: {e}", "Error")


def createPMCFolder(name):
    """Creacion de Carpeta PMC"""
    folder_path = os.path.join(ruta_alta, name)
    try:
        os.makedirs(folder_path, exist_ok=True)  # exist_ok=True evita error si ya existe
        print(f"Carpeta {name} creada con éxito")
    except Exception as e:
        Messagebox.show_error(f"Error al crear carpeta: {e}", "Error")


def refreshLogFile(data):
    """Actualiza el log de altas PMC"""
    try:
        with open("utils/log.txt", "a") as f:
            f.write(data + "\n")
        print("Log actualizado")
        return True
    except Exception as e:
        Messagebox.show_error(f"Error al escribir en el log: {e}", "Error")
        return False

      
def searchPMCCode(ruta):
    """Busca el codigo PMC en el archivo .txt"""
    list_codigo = []
    try:
        with open(ruta, "r") as f:
            for line in f:
                list_codigo.append(line.strip())
        print(f"Codigos PMC encontrados: {list_codigo}")
        return list_codigo
    except FileNotFoundError:
        Messagebox.show_error("Archivo codigosSIM.txt no encontrado", "Error")
        return None
    

async def createCSVFile(hardware_name, pmc_name, pmc_code, last_codigo_pmc):
    """Crea el archivo CSV del PMC"""
    """Códigos de referencia para alarmas PC"""
    ref_codigo_pmc = 997
    ref_codigo_doble_pmc = 999
    ref_codigo_alarma = 157
    
    codigo_alarma_PMC_PC = '"' + str(ref_codigo_pmc) + '"'
    codigo_alarma_PMC = '"' + str(ref_codigo_doble_pmc) + '"'
    codigo_alarma = '"' + str(ref_codigo_alarma) + ':'
    print(f"Código de alarma PMC: {codigo_alarma_PMC}")
    print(f"Código de alarma PMC PC: {codigo_alarma_PMC_PC}")
 
    
  
    csv_filename = f"{pmc_code}.csv"
    
    try:
        # Copiar archivo de referencia
        shutil.copyfile(f"utils/{hardware_name}.csv", csv_filename)
        
        # Leer contenido
        with open(csv_filename, "r", encoding="utf-8") as f:
            contenido = f.read()
        
        # Hacer reemplazos
        contenido = contenido.replace(hardware_name, pmc_name)
        contenido = contenido.replace(pmc_name, pmc_code)  # Segunda pasada
        contenido = contenido.replace(pmc_code + " ", pmc_name +" ") # Tercera pasada para asegurar reemplazo completo
        
        # Actualizar códigos
        nuevo_codigo_alarmas_PMC =  str(last_codigo_pmc[0])
        nuevo_codigo_alarmas_PMC_PC =  str(int(last_codigo_pmc[0])+2)
        nuevo_codigo_alarmas = str(int(last_codigo_pmc[1])+1)
        
        print(f"Nuevo código de alarma PMC: {nuevo_codigo_alarmas_PMC}")
        print(f"Nuevo código de alarma PMC PC: {nuevo_codigo_alarmas_PMC_PC}")
        print(f"Nuevo código PMC: {nuevo_codigo_alarmas}")
        
        # Reemplazos de códigos de alarma PC
        for i in range(2):
            contenido = contenido.replace(f'"{str(int(ref_codigo_pmc)+i)}"', f'"{(str(int(nuevo_codigo_alarmas_PMC)+i))}"')
            
        contenido = contenido.replace(str(codigo_alarma_PMC), f'"{(str(int(nuevo_codigo_alarmas_PMC_PC)))}"')
        
        print(f"Reemplazando código de alarma PMC: {codigo_alarma} por {nuevo_codigo_alarmas}") 
        contenido = contenido.replace(codigo_alarma, f'"{(str(nuevo_codigo_alarmas))}:')
        
        
        # Escribir contenido modificado
        with open(csv_filename, "w", encoding="utf-8") as f:
            f.write(contenido)
        

        # Actualizar códigos en archivo de texto
        with open("utils/codigosSIM.txt", "w") as f:
            f.write(f"{int(last_codigo_pmc[0]) + 3}\n{int(last_codigo_pmc[1]) + 1}")
        
        
        # AL FINALIZAR LA CREACION DEL CSV, SE MUEVEN LOS ARCHIVOS A LA CARPETA CORRESPONDIENTE
        # Mover archivos
        moveFiles(pmc_code, pmc_name)
        
        # Log de éxito
        refreshLogFile(
            f"PMC: {pmc_name} - Codigo: {pmc_code} - Hardware: {hardware_name} - "
            f"Fecha: {time.strftime('%Y-%m-%d %H:%M:%S')} - CREADO con Éxito"
        )
        print(f"✓ Archivo CSV creado: {csv_filename}")
        
    except FileNotFoundError:
        Messagebox.show_error(f"Archivo de referencia no encontrado: utils/{hardware_name}.csv", "Error")
        return
    except Exception as e:
        Messagebox.show_error(f"Error al crear CSV: {e}", "Error")
        refreshLogFile(
            f"PMC: {pmc_name} - Codigo: {pmc_code} - Hardware: {hardware_name} - "
            f"Fecha: {time.strftime('%Y-%m-%d %H:%M:%S')} - ERROR al dar de Alta el PMC"
        )
        return


async def createTGDFile(pmc_code):
    """Crea el archivo TGD del PMC"""
    tgd_filename = f"OSEDIS_{pmc_code}.xlsx"
    
    try:
        # Copiar archivo de referencia
        shutil.copyfile("utils/TagGrupo.xlsx", tgd_filename)
        
        # Leer contenido xlsx
        tagGroup = openpyxl.load_workbook(tgd_filename)
        sheet = tagGroup.active
        
        # Reemplazar RTU por el código PMC
        for row in range(1, sheet.max_row+1):
            cell = sheet.cell(row=row, column=2)
            if isinstance(cell.value, str):
                print(f"Revisando celda B{row}: {cell.value}")
                cell.value = re.sub("RTU", pmc_code,cell.value,flags=re.IGNORECASE)
                #texto = re.sub("RTU", pmc_code,cell.value,flags=re.IGNORECASE)
                #cell.value = texto
                #print(f"Nueva celda B{row}: {cell.value}")
                
        openpyxl.writer.excel.save_workbook(tagGroup, tgd_filename)    

        
        print(f"✓ Archivo Xlsx creado: {tgd_filename}")
        # Cambio de Extensión a .tgd
        os.rename(tgd_filename, f"OSEDIS_{pmc_code}.tgd")
        print(f"✓ Archivo TGD renombrado: OSEDIS_{pmc_code}.tgd")
        
    except FileNotFoundError:
        Messagebox.show_error("Archivo TagGrupo.xlsx no encontrado", "Error")
        return
    except Exception as e:
        Messagebox.show_error(f"Error al crear TGD: {e}", "Error")
        return


def moveFiles(pmc_code, pmc_name):
    """Mueve los archivos creados a la carpeta del PMC"""
    try:
        destination_folder = os.path.join(ruta_alta, pmc_name)
        os.makedirs(destination_folder, exist_ok=True)
        
        csv_source = f"{pmc_code}.csv"
        tgd_source = f"OSEDIS_{pmc_code}.tgd"
        
        csv_dest = os.path.join(destination_folder, f"{pmc_code}.csv")
        tgd_dest = os.path.join(destination_folder, f"OSEDIS_{pmc_code}.tgd")
        
        if os.path.exists(csv_source):
            shutil.move(csv_source, csv_dest)
            print(f"✓ {csv_source} movido a {destination_folder}")
        
        if os.path.exists(tgd_source):
            shutil.move(tgd_source, tgd_dest)
            print(f"✓ {tgd_source} movido a {destination_folder}")
        
    except Exception as e:
        Messagebox.show_error(f"Error al mover archivos: {e}", "Error")


if __name__ == "__main__":
    root = ttk.Window(themename="superhero")
    app = App(root)
    root.mainloop()